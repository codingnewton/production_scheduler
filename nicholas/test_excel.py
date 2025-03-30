from openpyxl import load_workbook
import pandas as pd
from scheduler import *
from datetime import time

workbook = load_workbook(filename=r"Test1_withMemberTaskExcel\Member&Task_Info.xlsx", data_only=True)
sheet = workbook['Members']
sheet2 = workbook['Tasks']

member_names = []
roles = []
member_hr_rates = []
member_ot_rates = []

for cell in sheet[3][1:]:
    if cell.value:
        member_names.append(cell.value)

        col_index = cell.col_idx
        roles.append(sheet.cell(row=4, column=col_index).value)
        member_hr_rates.append(sheet.cell(row=5, column=col_index).value)
        member_ot_rates.append(sheet.cell(row=6, column=col_index).value)

member_roles = [key for string in roles for key, value in ROLE_NAME.items() if value == string]

def format_time(row_index):
    hours = row_index // 4
    minutes = (row_index % 4) * 15
    if hours == 24 and minutes == 0:
        return 23, 59
    else:
        return hours, minutes

def process_unavailability(file_path, sheet_name="Members_availability"):
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    times = df.iloc[:, 0]
    results = {}
    for col_idx in range(1, df.shape[1]): 
        availability = df.iloc[:, col_idx].values 
        intervals = []
        start = None
        for i, available in enumerate(availability):
            if available == 0 and start is None:
                start = i 
            elif available == 1 and start is not None:
                intervals.append((format_time(start-1), format_time(i-1)))
                start = None
        if start is not None:
            intervals.append((format_time(start-1), format_time(len(availability)-1)))
        results[f"Person_{col_idx}"] = intervals
    return results

file_path = r"Test1_withMemberTaskExcel\Member&Task_Info.xlsx"
sheet_name = "Members_availability"
unavailability_results = process_unavailability(file_path, sheet_name)
member_intervals = []
for person, intervals in unavailability_results.items():
    member_intervals.append(intervals)

for i, name in enumerate(member_names):
    converted_intervals = [
        f"(time({start[0]}, {start[1]}), time({end[0]}, {end[1]}))" for start, end in member_intervals[i]]
    converted_intervals_str = f"[{', '.join(converted_intervals)}]"
    exec(f"{name} = Member({i}, '{name}', member_hr_rates[{i}], member_ot_rates[{i}], member_roles[{i}], blocked_timeslots={converted_intervals_str})")

members = [globals()[name] for name in member_names]

task_names = []
task_locations = []
task_durations = []
task_dependencies = []
# task_timeofday = []
task_actors = []
task_roles = []

for row in range(3, sheet2.max_row+1):
    cell_value = sheet2.cell(row=row, column=3).value
    if cell_value:
        task_names.append(cell_value)
        task_locations.append(sheet2.cell(row=row, column = 9).value)
        task_durations.append(sheet2.cell(row=row, column = 10).value)
        task_dependencies.append(sheet2.cell(row=row, column = 12).value)
        task_actors.append(sheet2.cell(row=row, column = 15).value)
        task_roles.append(sheet2.cell(row=row, column = 24).value)

for i, name in enumerate(task_names):
    task_location = task_locations[i]
    task_duration = task_durations[i]
    task_actor = task_actors[i]
    task_role = task_roles[i]
    task_dependency = task_dependencies[i]
    exec(f"{name} = Task({i}, {task_location}, {task_duration}, members={task_actor}, roles={task_role}, dependencies={task_dependency})")

tasks = [globals()[name] for name in task_names]

model = Optimizer(tasks, members)
res = model.schedule_tasks(timeout=20)
print('The optimized cost is: ', model.optimized_cost)
model.export_schedule(r'C:\Users\nicho\OneDrive - HKUST Connect\Study\FYP\output')
# export_schedules(res, model, location=r'C:\Users\nicho\OneDrive - HKUST Connect\Study\FYP\output')